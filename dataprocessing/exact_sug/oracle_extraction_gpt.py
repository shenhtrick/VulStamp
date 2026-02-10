from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os

# 读取CSV文件
df = pd.read_excel('.xlsx')

# 设置OpenAI客户端
client = OpenAI(
    base_url='',
    api_key='sk-***',
)

def analyze_code(code):
    messages = [
        {
            "role": "user",
            "content": "Given the vulnerable function, briefly suggest how to fix it." + code
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model="gpt-3.5-turbo",
        temperature=0.1,
    )
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            result['oracle_gpt'] = line
    return result

results_df = pd.DataFrame(columns=['func_before', 'oracle_gpt'])
count = 0
start_index = 0
excel_filename = '.xlsx'

for index, row in df.iloc[start_index:].iterrows():
    result = analyze_code(row['func_before'])
    retries = 0
    new_row = [
        row['func_before'],
        result.get('oracle_gpt', ''),
    ]
    if not os.path.exists(excel_filename):
        wb = Workbook()
        ws = wb.active
        ws.append(['func_before', 'oracle_gpt'])
        ws.append(new_row)
        wb.save(excel_filename)
    else:
        wb = load_workbook(excel_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(excel_filename)
    count += 1
    print("do it over {}".format(count))